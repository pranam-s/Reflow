"""Tests for reflow.taxonomy.reasons."""

from pathlib import Path

import openpyxl
import pytest

from reflow.taxonomy.provenance import resolve_vendored_path
from reflow.taxonomy.reasons import ReasonSpreadsheetError, parse_reason_records

REPO_ROOT = Path(__file__).resolve().parents[2]
VENDORED_PATH = resolve_vendored_path(REPO_ROOT)


def test_parse_reason_records_returns_114_records() -> None:
    records = parse_reason_records(VENDORED_PATH)
    assert len(records) == 114


def test_parse_reason_records_preserves_file_order_and_row_index() -> None:
    records = parse_reason_records(VENDORED_PATH)
    for expected_index, record in enumerate(records):
        assert record.row_index == expected_index


def test_parse_reason_records_preserves_duplicate_reason_codes() -> None:
    records = parse_reason_records(VENDORED_PATH)
    reasons = [record.reason for record in records]
    assert reasons.count("funds_blocked_by_mandate") == 2
    assert reasons.count("psp_not_available") == 2
    assert reasons.count("issuer_technical_error") == 2
    assert reasons.count("payment_method_not_enabled") == 2


def test_parse_reason_records_keeps_vendor_typo_verbatim() -> None:
    records = parse_reason_records(VENDORED_PATH)
    reasons = {record.reason for record in records}
    assert "psp_app_ not_available" in reasons


def test_first_record_matches_known_content() -> None:
    records = parse_reason_records(VENDORED_PATH)
    first = records[0]
    assert first.reason == "amount_less_than_minimum_amount"
    assert "minimum amount" in first.explanation
    assert "minimum fees" in first.next_steps


def test_parse_reason_records_rejects_bad_header(tmp_path: Path) -> None:
    bad_path = tmp_path / "bad_header.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.append(["Reason", "Wrong Column", "Next Steps"])
    worksheet.append(["some_reason", "an explanation", "do something"])
    workbook.save(bad_path)

    with pytest.raises(ReasonSpreadsheetError, match="Unexpected header row"):
        parse_reason_records(bad_path)


def test_parse_reason_records_rejects_empty_workbook(tmp_path: Path) -> None:
    empty_path = tmp_path / "empty.xlsx"
    workbook = openpyxl.Workbook()
    workbook.save(empty_path)

    with pytest.raises(ReasonSpreadsheetError, match="no header row"):
        parse_reason_records(empty_path)


def test_parse_reason_records_rejects_missing_next_steps(tmp_path: Path) -> None:
    bad_path = tmp_path / "missing_next_steps.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.append(["Reason", "Explanation", "Next Steps"])
    worksheet.append(["some_reason", "an explanation", None])
    workbook.save(bad_path)

    with pytest.raises(ReasonSpreadsheetError, match="missing an explanation or next-steps"):
        parse_reason_records(bad_path)


def test_parse_reason_records_rejects_wrong_row_count(tmp_path: Path) -> None:
    short_path = tmp_path / "short.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.append(["Reason", "Explanation", "Next Steps"])
    worksheet.append(["some_reason", "an explanation", "do something"])
    workbook.save(short_path)

    with pytest.raises(ReasonSpreadsheetError, match="Parsed 1 data rows, expected 114"):
        parse_reason_records(short_path)
